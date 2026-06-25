"""
parsers.py - Log format parsers for log-analyzer.
Auto-detects and parses common log formats including:
- ISO-8601 timestamp logs (Java-style)
- Syslog format
- NCSA Common Log Format (Apache/Nginx)
- Java stack traces
- Java application logs with thread names
- Custom formats
"""

import re
from datetime import datetime
from typing import Iterator, Optional, Dict, Any, List, Tuple


LOG_LEVELS = {"TRACE", "DEBUG", "INFO", "WARN", "WARNING", "ERROR", "FATAL", "CRITICAL"}


FORMAT_PATTERNS: List[Tuple[str, re.Pattern, Dict[str, Any]]] = [
    ("java_app", re.compile(
        r'^(?P<ts>\d{4}[-/]\d{2}[-/]\d{2}[T ]\d{2}:\d{2}:\d{2}(?:[.,]\d{3})?)'
        r'\s+\[(?P<thread>[^\]]+)\]'
        r'\s+(?P<level>TRACE|DEBUG|INFO|WARN(?:ING)?|ERROR|FATAL|CRITICAL)\b'
        r'\s+(?P<trace_id>[a-fA-F0-9-]+)?\s*'
        r'(?P<module>[\w.]+\.[\w.]+)?'
        r'\s*[:\-]?\s*(?P<message>.+)?',
        re.IGNORECASE
    ), {}),
    ("iso", re.compile(
        r'^\[?(?P<ts>\d{4}[-/]\d{2}[-/]\d{2}[T ]\d{2}:\d{2}:\d{2}(?:[.,]\d{3})?)\]?\s+'
        r'(?P<level>TRACE|DEBUG|INFO|WARN(?:ING)?|ERROR|FATAL|CRITICAL)\b'
        r'(?:\s+\[?(?P<module>[^\]]+)\]?)?'
        r'\s*[:\-]?\s*(?P<message>.+)',
        re.IGNORECASE
    ), {}),
    ("syslog", re.compile(
        r'^(?:<\d+>)?'
        r'(?P<ts>[A-Z][a-z]{2}\s+\d{1,2}\s+\d{2}:\d{2}:\d{2}(?:\.\d+)?)'
        r'\s+\S+\s+(?P<module>\S+?)(?:\[\d+\])?:\s*'
        r'(?P<level>TRACE|DEBUG|INFO|WARN(?:ING)?|ERROR|FATAL)\s*'
        r'[:\-]?\s*(?P<message>.+)?',
        re.IGNORECASE
    ), {}),
    ("ncsa", re.compile(
        r'^(?P<remote>\S+)\s+\S+\s+\S+\s+'
        r'\[(?P<ts>\d{2}/\w{3}/\d{4}:\d{2}:\d{2}:\d{2}\s[+-]\d{4})\]'
        r'\s+"(?P<method>\S+)\s+(?P<path>\S+)\s+\S+"\s+'
        r'(?P<status>\d{3})\s+(?P<size>\d+|-)',
    ), {}),
    ("java_stack", re.compile(
        r'^\s+at\s+[\w.$]+\([\w.]+\.java:\d+\)',
    ), {}),
    ("caused_by", re.compile(
        r'^Caused by:\s+(?P<exception>[\w.]+(?:Exception|Error)):\s*(?P<message>.*)',
    ), {}),
    ("json_log", re.compile(
        r'^\s*\{.*"level"\s*:\s*"(?P<level>[A-Z]+)".*\}',
        re.DOTALL
    ), {}),
]


def detect_format(line: str) -> Optional[str]:
    """Auto-detect which log format a line belongs to."""
    if not line.strip():
        return None
    for name, pattern, _ in FORMAT_PATTERNS:
        if pattern.match(line):
            return name
    return None


def parse_line(line: str) -> Optional[Dict[str, Any]]:
    """
    Parse a single log line into a structured record.
    Returns None for unrecognized lines (continuation lines, blank lines, etc.)
    """
    stripped = line.strip("\n\r")
    if not stripped:
        return None

    for name, pattern, _ in FORMAT_PATTERNS:
        m = pattern.match(stripped)
        if not m:
            continue

        record: Dict[str, Any] = {
            "_raw": line,
            "_format": name,
            "_line_no": 0,
        }
        gd = m.groupdict()

        level_raw = gd.get("level", "")
        if level_raw:
            record["level"] = level_raw.upper()
        elif name == "ncsa":
            status = int(gd.get("status", 0))
            if status >= 500:
                record["level"] = "ERROR"
            elif status >= 400:
                record["level"] = "WARN"
            else:
                record["level"] = "INFO"

        ts_raw = gd.get("ts")
        if ts_raw:
            record["timestamp"] = _parse_timestamp(ts_raw, name)
            record["timestamp_raw"] = ts_raw

        msg = gd.get("message", "").strip()
        if msg:
            record["message"] = msg

        mod = gd.get("module")
        if mod:
            record["module"] = mod.strip("[] ")

        thread = gd.get("thread")
        if thread:
            record["thread"] = thread

        trace_id = gd.get("trace_id")
        if trace_id:
            record["trace_id"] = trace_id

        if name == "ncsa":
            record["http_method"] = gd.get("method", "")
            record["http_path"] = gd.get("path", "")
            record["http_status"] = int(gd.get("status", 0))
            record["remote_addr"] = gd.get("remote", "")
            record["message"] = f'{gd.get("method","")} {gd.get("path","")} -> {gd.get("status","")}'

        if name == "caused_by":
            record["exception"] = gd.get("exception", "")
            record["is_root_cause"] = True

        if name == "java_stack":
            record["is_stack_trace"] = True

        if name in ("java_app", "iso") and record.get("level") in ("ERROR", "FATAL", "CRITICAL"):
            _extract_error_info(record)

        return record

    return None


def _extract_error_info(record: Dict[str, Any]) -> None:
    """Extract error type and error message from ERROR/FATAL/Critical level records."""
    msg = record.get("message", "")
    error_pattern = re.compile(r'([\w\.]+Exception|[\w\.]+Error):\s*(.*)$')
    ex_match = error_pattern.search(msg)
    if ex_match:
        record["error_type"] = ex_match.group(1)
        record["error_message"] = ex_match.group(2)


def _parse_timestamp(raw: str, fmt_name: str) -> Optional[str]:
    """Parse timestamp into ISO-8601 string."""
    try:
        if fmt_name in ("java_app", "iso"):
            ts = raw.replace("/", "-").replace("T", " ").replace(",", ".")
            if "." in ts:
                datetime.strptime(ts[:26], "%Y-%m-%d %H:%M:%S.%f")
            else:
                datetime.strptime(ts[:19], "%Y-%m-%d %H:%M:%S")
            return ts[:26] if "." in ts else ts[:19]

        elif fmt_name == "syslog":
            return datetime.strptime(raw, "%b %d %H:%M:%S").replace(year=datetime.now().year).isoformat()

        elif fmt_name == "ncsa":
            return datetime.strptime(raw[:20], "%d/%b/%Y:%H:%M:%S").isoformat()
    except (ValueError, IndexError):
        pass
    return None


def is_stack_trace_line(line: str) -> bool:
    """Check if a line is a Java stack continuation line."""
    return bool(re.match(r'^\s+at\s+', line)) or bool(re.match(r'^\.\.\.\s+\d+\s+more', line))


def parse_json_log(line: str) -> Optional[Dict[str, Any]]:
    """Parse JSON-formatted log lines."""
    import json
    stripped = line.strip()
    if not stripped.startswith("{"):
        return None
    try:
        data = json.loads(stripped)
        record: Dict[str, Any] = {
            "_raw": line,
            "_format": "json_log",
            "_line_no": 0,
        }
        if "level" in data:
            record["level"] = str(data["level"]).upper()
        if "message" in data:
            record["message"] = str(data["message"])
        if "timestamp" in data or "time" in data:
            ts_key = "timestamp" if "timestamp" in data else "time"
            record["timestamp_raw"] = str(data[ts_key])
            record["timestamp"] = _normalize_json_timestamp(data[ts_key])
        if "module" in data or "logger" in data or "service" in data:
            for key in ["module", "logger", "service"]:
                if key in data:
                    record["module"] = str(data[key])
                    break
        if "error" in data:
            record["error_type"] = str(data["error"].get("type", "")) if isinstance(data["error"], dict) else str(data["error"])
        if "stack" in data:
            record["stack_trace"] = data["stack"]
        return record
    except json.JSONDecodeError:
        return None


def _normalize_json_timestamp(ts: Any) -> Optional[str]:
    """Normalize various JSON timestamp formats to ISO-8601."""
    ts_str = str(ts)
    try:
        if ts_str.isdigit():
            ts_int = int(ts_str)
            if ts_int > 10**12:
                return datetime.fromtimestamp(ts_int / 1000).isoformat()
            return datetime.fromtimestamp(ts_int).isoformat()
        for fmt in ("%Y-%m-%dT%H:%M:%S.%fZ", "%Y-%m-%dT%H:%M:%SZ",
                    "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(ts_str, fmt).isoformat()
            except ValueError:
                continue
    except (ValueError, TypeError):
        pass
    return None


def parse_excel_file(file_path: str) -> Iterator[Dict[str, Any]]:
    """Parse Excel (.xlsx) log files."""
    try:
        import openpyxl
    except ImportError:
        raise ImportError("openpyxl is required for Excel parsing. Install with: pip install openpyxl")

    wb = openpyxl.load_workbook(file_path, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(cell) if cell is not None else f"col_{i}" for i, cell in enumerate(row)]
                continue

            record = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = cell

            # Convert to log record format
            yield _convert_to_log_record(record, "excel")


def parse_csv_file(file_path: str) -> Iterator[Dict[str, Any]]:
    """Parse CSV log files."""
    import csv
    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield _convert_to_log_record(row, "csv")


def parse_xml_file(file_path: str) -> Iterator[Dict[str, Any]]:
    """Parse XML log files."""
    import xml.etree.ElementTree as ET
    tree = ET.parse(file_path)
    root = tree.getroot()

    # Try to find log entries (common patterns)
    log_entries = root.findall('.//log') or root.findall('.//entry') or root.findall('.//record') or root.findall('.//event')

    if not log_entries:
        # If no standard log tags, treat each child element as a log entry
        log_entries = list(root)

    for entry in log_entries:
        record = {}
        for child in entry:
            record[child.tag] = child.text

        yield _convert_to_log_record(record, "xml")


def parse_pdf_file(file_path: str) -> Iterator[Dict[str, Any]]:
    """Parse PDF log files (extract text and parse as lines)."""
    try:
        import PyPDF2
    except ImportError:
        raise ImportError("PyPDF2 is required for PDF parsing. Install with: pip install PyPDF2")

    with open(file_path, 'rb') as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            text = page.extract_text()
            for line in text.split('\n'):
                if line.strip():
                    record = parse_line(line)
                    if record:
                        record["_format"] = "pdf"
                        yield record


def parse_json_file(file_path: str) -> Iterator[Dict[str, Any]]:
    """Parse JSON log files (line-delimited JSON or JSON array)."""
    import json

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
        content = f.read().strip()

        # Try to parse as JSON array first
        try:
            data = json.loads(content)
            if isinstance(data, list):
                for item in data:
                    if isinstance(item, dict):
                        yield _convert_to_log_record(item, "json_file")
                return
        except json.JSONDecodeError:
            pass

        # Fall back to line-delimited JSON
        for line in content.split('\n'):
            if line.strip():
                record = parse_json_log(line)
                if record:
                    yield record


def _convert_to_log_record(data: Dict[str, Any], format_type: str) -> Dict[str, Any]:
    """Convert parsed data to standard log record format."""
    record: Dict[str, Any] = {
        "_raw": str(data),
        "_format": format_type,
        "_line_no": 0,
    }

    # Map common field names
    field_mapping = {
        'level': ['level', 'severity', 'priority', 'log_level'],
        'message': ['message', 'msg', 'text', 'description', 'content'],
        'timestamp': ['timestamp', 'time', 'date', '@timestamp', 'datetime'],
        'module': ['module', 'logger', 'service', 'source', 'component', 'app'],
        'error_type': ['error_type', 'exception', 'error', 'exception_type'],
        'thread': ['thread', 'thread_name', 'threadId'],
        'trace_id': ['trace_id', 'traceId', 'request_id', 'requestId'],
    }

    for target_field, source_fields in field_mapping.items():
        for source_field in source_fields:
            if source_field in data and data[source_field]:
                value = data[source_field]
                if target_field == 'level':
                    record[target_field] = str(value).upper()
                elif target_field == 'timestamp':
                    record['timestamp_raw'] = str(value)
                    record['timestamp'] = _normalize_json_timestamp(value)
                else:
                    record[target_field] = str(value)
                break

    # If no level found, try to infer from content
    if 'level' not in record:
        message = record.get('message', '').upper()
        for level in ['CRITICAL', 'FATAL', 'ERROR', 'WARN', 'WARNING', 'INFO', 'DEBUG', 'TRACE']:
            if level in message:
                record['level'] = level
                break

    # If still no level, default to INFO
    if 'level' not in record:
        record['level'] = 'INFO'

    # Ensure message field exists
    if 'message' not in record:
        record['message'] = str(data)

    return record
